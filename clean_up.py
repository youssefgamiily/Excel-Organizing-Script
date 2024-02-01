import openpyxl
import json

class Item:
    def __init__(self, inventory_code, auction_code, name, selling_price, wrapping_cost, status):
        self.inventory_code = inventory_code.value
        self.auction_code = auction_code
        self.name = name.value
        self.selling_price = selling_price.value
        self.wrapping_cost = wrapping_cost.value
        self.status = status.value

    def to_dict(self):
        return {
            "inventory_code": self.inventory_code,
            "auction_code": self.auction_code,
            "name": self.name,
            "selling_price": self.selling_price,
            "wrapping_cost": self.wrapping_cost,
            "status": self.status
        }
    def to_json(self):
        return json.dumps(self.to_dict(), indent=2)


class Buyer:
    def __init__(self, name, city, address, number, status):
        self.name = name.value
        self.city = city.value
        self.address = address.value
        self.number = number.value
        self.status = status.value

    def to_dict(self):
        return {
            "name": self.name,
            "city": self.city,
            "address": self.address,
            "number": self.number,
            "status": self.status
        }

print("hi")
wb_auctions = openpyxl.load_workbook('./Auctions Sales Recording 11 Oct 23.xlsx')
print(wb_auctions)
print(wb_auctions.sheetnames)

buyer_items_dict = {}  # Dictionary to store buyer objects and their associated items

for index, sheet in enumerate(wb_auctions.sheetnames):
    split_words = sheet.split()
    isAuction = split_words[0].isnumeric()
    auction_num = split_words[0] if isAuction else 9999
    print("int(isAuction) is: ", int(isAuction))
    if isAuction and int(auction_num) > 1759:
        sheet_name = wb_auctions.sheetnames[index]

        for row in wb_auctions[sheet_name].iter_rows(min_row=2, max_row=100, min_col=1, max_col=11):
            # class Item:
            # def __init__(self, inventory_code0, auction_code1, name2, selling_price3, wrapping_cost8, status):
            item = Item(row[0], row[1], row[2], row[3], row[8], row[9])
            print("item is: ", item.to_json)
            #         class Buyer:
            # def __init__(self, name4, city5, address6, number7, status):
            buyer = Buyer(row[3], row[4], row[5], row[6], row[10])

            # Write data directly to the dictionary
            if buyer.name in buyer_items_dict:
                buyer_items_dict[buyer.name][1].append(item.to_dict())
            else:
                buyer_items_dict[buyer.name] = [buyer.to_dict(), [item.to_dict()]]

            # Accessing the values for each cell in the row
            for cell in row:
                print(f"Sheet: {sheet_name}, Row: {cell.row}, Column: {cell.column}, Value: {cell.value}")

# Dump the dictionary into a JSON file
json_filename = "buyer_items_data.json"

with open(json_filename, mode='w') as json_file:
    json.dump(buyer_items_dict, json_file, indent=2)

print(f"Data has been dumped into '{json_filename}'.")
